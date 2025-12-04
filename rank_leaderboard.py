import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def clean_value(val):
    if pd.isna(val) or val == '-' or val == 'D$Q':
        return 0
    return float(val)

def get_countback_key(row, round_cols):
    scores = []
    for col in round_cols:
        scores.append(clean_value(row[col]))
    # Sort scores descending
    scores.sort(reverse=True)
    return scores

def rank_leaderboard():
    input_file = 'leaderboard.xlsx'
    output_file = 'leaderboard_sorted.xlsx'

    # Load the entire sheet to find tables
    df_all = pd.read_excel(input_file, sheet_name='Leaderboard', header=None)

    # Find the start of the top table (Leaderboard)
    # Looking for "Pos" and "Player" in the same row
    top_header_idx = -1
    for idx, row in df_all.iterrows():
        if 'Pos' in str(row.values) and 'Player' in str(row.values):
            top_header_idx = idx
            break
    
    if top_header_idx == -1:
        print("Could not find Leaderboard table.")
        return

    # Find the start of the bottom table (Spending)
    bottom_header_idx = -1
    for idx, row in df_all.iterrows():
        if idx > top_header_idx and 'Pos' in str(row.values) and 'Player' in str(row.values):
            bottom_header_idx = idx
            break

    if bottom_header_idx == -1:
        print("Could not find Spending table.")
        return

    # Extract Top Table
    df_top = df_all.iloc[top_header_idx:].copy()
    # We need to handle duplicate column names (Pts, Pts...)
    # Let's manually assign meaningful names based on inspection
    # Col 0: Pos, Col 1: Player, Col 2-25: R01-R24, Col 26: Points, Col 27: Spent ($m), Col 28: $m/Pt
    
    new_columns = ['Pos', 'Player']
    # There are 24 rounds (R01 to R24) based on inspection (cols 2 to 25)
    # Let's verify the number of columns
    num_cols = df_top.shape[1]
    # Assuming standard structure
    for i in range(1, 25):
        new_columns.append(f'R{i:02d}')
    
    # Remaining columns
    # We expect Total Points, Spent, CostPerPt
    # But let's just append generic names for the rest if they don't match
    # The inspection shows 29 columns total.
    # 2 + 24 = 26. So 26, 27, 28 remain.
    new_columns.extend(['Points', 'Spent ($m)', '$m/Pt'])
    
    # Check if length matches
    if len(new_columns) != num_cols:
        # Fallback: just deduplicate existing names
        cols = list(df_top.iloc[0])
        counts = {}
        new_columns = []
        for c in cols:
            if c not in counts:
                counts[c] = 0
                new_columns.append(c)
            else:
                counts[c] += 1
                new_columns.append(f"{c}.{counts[c]}")
    
    df_top.columns = new_columns
    df_top = df_top.iloc[1:]
    
    # Identify the end of the player list. 
    last_player_idx_top = -1
    for i in range(len(df_top)):
        val = df_top.iloc[i]['Player']
        if str(val) == 'Points Totals' or pd.isna(val):
            last_player_idx_top = i
            break
    
    if last_player_idx_top != -1:
        df_players_top = df_top.iloc[:last_player_idx_top].copy()
    else:
        df_players_top = df_top.copy()

    # Extract Bottom Table
    df_bottom = df_all.iloc[bottom_header_idx:].copy()
    df_bottom.columns = df_bottom.iloc[0]
    df_bottom = df_bottom.iloc[1:]
    
    last_player_idx_bottom = -1
    for i in range(len(df_bottom)):
        val = df_bottom.iloc[i]['Player']
        if str(val) == 'Spending Totals' or pd.isna(val):
            last_player_idx_bottom = i
            break
            
    if last_player_idx_bottom != -1:
        df_players_bottom = df_bottom.iloc[:last_player_idx_bottom].copy()
    else:
        df_players_bottom = df_bottom.copy()

    # Columns for rounds: R01, R02...
    # If we used the manual renaming, they are R01...R24
    round_cols = [c for c in df_players_top.columns if str(c).startswith('R') and len(str(c)) == 3 and c[1:].isdigit()]
    
    # Calculate Sort Keys
    def create_sort_key(row):
        # Use the column names we assigned
        # If manual renaming worked, we use 'Points' and 'Spent ($m)'
        # If fallback used, we might have 'Points' and 'Spent ($m)' if they were unique in header
        
        # Check if 'Points' exists
        pts_col = 'Points' if 'Points' in row else 'Total Points' # Fallback
        if pts_col not in row:
             # Try to find column with 'Points' or 'Pts' that is NOT a round column
             pass
        
        pts = clean_value(row.get(pts_col, 0))
        spend = clean_value(row.get('Spent ($m)', 0))
        
        countback = get_countback_key(row, round_cols)
        neg_countback = tuple([-x for x in countback])
        
        name = str(row.get('Player', ''))
        
        return (-pts, spend, neg_countback, name)

    # Add sort key
    df_players_top['sort_key'] = df_players_top.apply(create_sort_key, axis=1)
    
    # Sort
    df_players_top_sorted = df_players_top.sort_values('sort_key')
    
    # Re-assign Ranks
    df_players_top_sorted['Pos'] = range(1, len(df_players_top_sorted) + 1)
    
    # Identify Ties for highlighting
    def create_tie_key(row):
        k = row['sort_key']
        return k[:-1] # Exclude name
    
    df_players_top_sorted['tie_key'] = df_players_top_sorted.apply(create_tie_key, axis=1)
    
    tied_keys = df_players_top_sorted[df_players_top_sorted.duplicated('tie_key', keep=False)]['tie_key'].unique()
    players_to_highlight = set(df_players_top_sorted[df_players_top_sorted['tie_key'].isin(tied_keys)]['Player'])
    
    # Reorder Bottom Table
    player_order = list(df_players_top_sorted['Player'])
    df_players_bottom['Player'] = df_players_bottom['Player'].astype(str)
    player_rank_map = {p: i for i, p in enumerate(player_order)}
    df_players_bottom['temp_rank'] = df_players_bottom['Player'].map(player_rank_map)
    df_players_bottom_sorted = df_players_bottom.sort_values('temp_rank')
    df_players_bottom_sorted['Pos'] = range(1, len(df_players_bottom_sorted) + 1)
    
    df_players_top_sorted = df_players_top_sorted.drop(columns=['sort_key', 'tie_key'])
    df_players_bottom_sorted = df_players_bottom_sorted.drop(columns=['temp_rank'])
    
    # --- Write to Excel ---
    wb = load_workbook(input_file)
    ws = wb['Leaderboard']
    
    start_row_top = top_header_idx + 2
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    
    for i, (idx, row) in enumerate(df_players_top_sorted.iterrows()):
        excel_row = start_row_top + i
        ws.cell(row=excel_row, column=1).value = i + 1
        
        for j in range(len(df_players_top_sorted.columns)):
            # Use iloc to get value by position, avoiding duplicate column name issues
            val = row.iloc[j]
            ws.cell(row=excel_row, column=j+1).value = val
            
            col_name = df_players_top_sorted.columns[j]
            if col_name == 'Player' and val in players_to_highlight:
                ws.cell(row=excel_row, column=j+1).fill = red_fill
            elif col_name == 'Player':
                 ws.cell(row=excel_row, column=j+1).fill = PatternFill(fill_type=None)

    # Update Bottom Table
    start_row_bottom = bottom_header_idx + 2
    for i, (idx, row) in enumerate(df_players_bottom_sorted.iterrows()):
        excel_row = start_row_bottom + i
        ws.cell(row=excel_row, column=1).value = i + 1
        for j in range(len(df_players_bottom_sorted.columns)):
            val = row.iloc[j]
            ws.cell(row=excel_row, column=j+1).value = val
            
            col_name = df_players_bottom_sorted.columns[j]
            if col_name == 'Player' and val in players_to_highlight:
                ws.cell(row=excel_row, column=j+1).fill = red_fill
            elif col_name == 'Player':
                 ws.cell(row=excel_row, column=j+1).fill = PatternFill(fill_type=None)

    wb.save(output_file)
    print(f"Sorted leaderboard saved to {output_file}")

if __name__ == "__main__":
    rank_leaderboard()
